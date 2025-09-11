import json
import os.path

import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl import load_workbook
from typing import List, Dict, Any, Optional

def save(results_list: List[Dict[str, Any]],
         out_json: Optional[str], out_excel: Optional[str]) -> None:
    """
    Сохраняет результаты работы парсера:
      - в JSON-файл (если указан out_json),
      - в Excel (если указан out_excel).

    В Excel добавляется оформление:
      - фиксированные ширины столбцов,
      - выравнивание текста по центру,
      - перенос по словам для длинных ячеек.
    """
    # Сохранение в JSON
    if out_json:
        try:
            os.makedirs(os.path.dirname(os.path.abspath(out_json)), exist_ok=True)
            with open(out_json, "w", encoding="utf-8") as f:
                json.dump(results_list, f, ensure_ascii=False, indent=2)
        except IOError as e:
            print(f"Error: Could not save JSON to {out_json}. Error: {e}")

    # Сохранение в Excel
    if out_excel:
        try:
            os.makedirs(os.path.dirname(os.path.abspath(out_excel)), exist_ok=True)
            df = pd.DataFrame(results_list)
            df.to_excel(out_excel, index=False)

            # Ширина столбцов (заранее подобранные значения для читаемости)
            col_widths = {
                "year": 6,
                "authors": 40,
                "title": 60,
                "doi": 30,
                "citations": 11,
                "impact_factor": 15,
                "quartile": 10,
                "link": 50,
                "available_on_site": 19,
                "researchgate": 14,
                "pirates": 9
            }

            wb = load_workbook(out_excel)
            ws = wb.active

            for i, col in enumerate(ws.columns, start=1):
                col_letter: str = get_column_letter(i)
                header: Optional[str] = ws.cell(row=1, column=i).value

                # Берём ширину из словаря или дефолтное значение
                width: int = col_widths.get(header, 15)
                ws.column_dimensions[col_letter].width = width

                # Выравнивание и перенос слов для всех ячеек
                for cell in col:
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            wb.save(out_excel)
        except IOError as e:
            print(f"Error: Could not save Excel to {out_excel}. Error: {e}")
